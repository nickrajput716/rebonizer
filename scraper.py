from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from webdriver_manager.chrome import ChromeDriverManager
from bs4 import BeautifulSoup
from datetime import datetime
from io import BytesIO
import time
import re

class SeleniumResultScraper:
    def __init__(self): 
        self.driver = None

    def setup_driver(self):
        opts = Options()
        opts.add_argument('--headless')
        opts.add_argument('--no-sandbox')
        opts.add_argument('--disable-dev-shm-usage')
        opts.add_argument('--disable-gpu')
        self.driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=opts)
        self.driver.implicitly_wait(10)

    def close_driver(self):
        if self.driver: 
            self.driver.quit(); 
            self.driver = None

    def get_clean_subject_code(self, raw):
        if not raw: return None
        code = raw.upper().strip()
        code = re.sub(r'[-_][A-Z0-9]*[A-Z]+$', '', code)
        code = re.sub(r'[A-Z0-9]*[A-Z]+$', '', code)
        code = code.replace('-', '').replace('_', '')
        return code if re.match(r'^[A-Z]{2,}\d{3,}$', code) else None

    def scrape_single_result(self, url, roll):
        try:
            self.driver.get(url); time.sleep(2)
            inp = None
            for sel in ["input[type='text']", "input[name*='roll']", "input[id*='roll']", ".form-control"]:
                els = self.driver.find_elements(By.CSS_SELECTOR, sel)
                if els: inp = els[0]; break
            if not inp: inp = self.driver.find_element(By.XPATH, "//input[@type='text']")
            inp.clear(); inp.send_keys(roll); time.sleep(0.5)

            btn = None
            for sel in ["input[type='submit']", "button[type='submit']", "input[value*='Search']", ".btn"]:
                els = self.driver.find_elements(By.CSS_SELECTOR, sel)
                if els: btn = els[0]; break
            if btn: btn.click()
            else: self.driver.find_element(By.XPATH, "//input[@type='submit']").click()
            time.sleep(3)
            soup = BeautifulSoup(self.driver.page_source, 'html.parser')
            return self.parse_result_page(soup, roll)
        except: 
            return {'roll_no': roll, 'name': 'ERROR', 'status': 'Error', 'subjects': {}, 'sgpa': 'N/A', 'supply_subjects': []}

    def parse_result_page(self, soup, roll):
        result = {'roll_no': roll, 'name': 'Not Found', 'father_name': 'N/A', 'status': 'Not Found',
                  'subjects': {}, 'supply_subjects': [], 'sgpa': 'N/A', 'cgpa': 'N/A'}
        text = soup.get_text().lower()
        if any(x in text for x in ['no record', 'invalid', 'not found']): return result

        for tr in soup.find_all('tr'):
            tds = tr.find_all(['td','th'])
            if len(tds) >= 2:
                label = tds[0].get_text(strip=True).lower()
                val = tds[1].get_text(strip=True)
                if 'name' in label and 'father' not in label: result['name'] = val
                if 'father' in label: result['father_name'] = val

        for table in soup.find_all('table'):
            rows = table.find_all('tr')
            if len(rows) < 2: continue
            header = " ".join(c.get_text().lower() for c in rows[0].find_all(['td','th']))
            if not any(k in header for k in ['subject', 'code', 'paper id']): continue

            for row in rows[1:]:
                cells = [c.get_text(strip=True) for c in row.find_all(['td','th'])]
                if len(cells) < 3: continue
                raw_code = cells[0].upper()
                name = cells[1]
                grade = cells[-1]
                credit = cells[2] if len(cells)>3 and re.match(r'^\d+\.?\d*$', cells[2]) else ''

                clean = self.get_clean_subject_code(raw_code)
                if not clean: continue

                if clean not in result['subjects']:
                    result['subjects'][clean] = {'name': name if name != raw_code else 'Unknown', 'credit': credit, 'grade': grade if grade not in ['-',''] else '–'}
                else:
                    cur = result['subjects'][clean]
                    if cur['name'] == 'Unknown': cur['name'] = name
                    if not cur['credit'] and credit: cur['credit'] = credit
                    if grade not in ['-','','–']: cur['grade'] = grade

                if grade and grade.upper() in ['F','FAIL']:
                    result['supply_subjects'].append(clean)

        full = soup.get_text()
        m = re.search(r'SGPA[:\s]*([0-9.]+)', full, re.I)
        if m: result['sgpa'] = m.group(1)
        m = re.search(r'CGPA[:\s]*([0-9.]+)', full, re.I)
        if m: result['cgpa'] = m.group(1)
        if 'pass' in text: result['status'] = 'PASS'
        elif 'fail' in text: result['status'] = 'FAIL'
        return result

    def bulk_scrape(self, url, start, end, delay):
        results = []
        self.setup_driver()
        try:
            for n in range(int(start), int(end)+1):
                print(f"[COBRA TECH] Scraping {n}")
                results.append(self.scrape_single_result(url, str(n)))
                if n < int(end): time.sleep(delay)
        finally:
            self.close_driver()
        return results

    def generate_excel(self, results):
        if not results: return None
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        all_codes = sorted({c for r in results for c in r['subjects']})
        meta = {}
        for r in results:
            for c, d in r['subjects'].items():
                if c not in meta or meta[c]['name'] == 'Unknown':
                    meta[c] = {'name': d['name'], 'credit': d['credit']}

        wb = Workbook()
        ws = wb.active
        ws.title = "COBRA TECH Results"

        header_fill = PatternFill(start_color="93BFC7", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        border = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))

        headers = ['Roll No.', 'Student Name', "Father's Name"]
        for c in all_codes:
            m = meta.get(c, {'name': 'Unknown', 'credit': ''})
            txt = f"{m['name']}\n({c})" + (f"\n[{m['credit']}]" if m['credit'] else "")
            headers.append(txt)
        headers.extend(['SGPA', 'CGPA', 'Result', 'Supply'])

        for col, h in enumerate(headers, 1):
            cell = ws.cell(1, col, h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center
            cell.border = border

        for idx, r in enumerate(results, 2):
            row = [r['roll_no'], r['name'] if r['name'] not in ['Not Found','ERROR'] else 'Not Found', r.get('father_name','N/A')]
            for c in all_codes:
                row.append(r['subjects'][c]['grade'] if c in r['subjects'] else '–')
            row.extend([r['sgpa'], r['cgpa'], r['status'], ', '.join(r['supply_subjects']) if r['supply_subjects'] else '–'])
            for col, val in enumerate(row, 1):
                cell = ws.cell(idx, col, val)
                cell.alignment = center
                cell.border = border

        ws.row_dimensions[1].height = 80
        ws.freeze_panes = 'A2'

        bio = BytesIO()
        wb.save(bio)
        bio.seek(0)
        return bio